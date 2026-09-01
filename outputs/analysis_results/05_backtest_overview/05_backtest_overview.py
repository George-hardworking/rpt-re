"""05 — backtest overview tables (baseline + top500, US + CN).

Read-only: loads step-04 ``cnn_baseline`` and ``cnn_top500`` H1 xlsx.
Writes one workbook with four presentation-style sheets (US and CN rows
combined; ``Market`` column distinguishes markets):

  baseline_presentation, baseline_deciles, top500_presentation, top500_deciles

Plus ``notes``, ``raw_baseline``, ``raw_top500``.

Run (from repo root, 5020_env):
  python outputs/analysis_results/05_backtest_overview/05_backtest_overview.py
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / "src"))

from config import (
    BACKTEST_CNN_ROOT,
    BACKTEST_CNN_TOP500_ROOT,
    MARKET_CN,
    MARKET_US,
)

HERE = Path(__file__).resolve().parent

PortfolioKind = Literal["baseline", "top500"]

PORTFOLIO_ROOTS: dict[PortfolioKind, Path] = {
    "baseline": BACKTEST_CNN_ROOT,
    "top500": BACKTEST_CNN_TOP500_ROOT,
}

MARKETS = (MARKET_US, MARKET_CN)

ACT_RET_METRIC = "Annualized Active Return"
IR_METRIC = "Information Ratio"
MDD_METRIC = "Max Drawdown (Active)"
TO_METRIC = "Turnover (annualized)"

_CONFIG_RE = re.compile(r"^(?P<tag>I\d+_R\d+)_h(?P<h>\d+)\.xlsx$", re.IGNORECASE)

PRESENTATION_HEADERS = [
    "Market",
    "Weight",
    "Freq",
    "Config",
    "Ret1",
    "IR",
    "MaxDD",
    "Turnover (ann.)",
    "TO (monthly %, GKX)",
    "D1",
    "D10",
]
PRESENTATION_FMT = {
    "Market": "text",
    "Weight": "text",
    "Freq": "text",
    "Config": "text",
    "Ret1": "pct",
    "IR": "ir",
    "MaxDD": "pct",
    "Turnover (ann.)": "turn",
    "TO (monthly %, GKX)": "turn",
    "D1": "pct",
    "D10": "pct",
}
PRESENTATION_FILLS = {
    "Market": "4472C4",
    "Weight": "4472C4",
    "Freq": "4472C4",
    "Config": "4472C4",
    "Ret1": "548235",
    "IR": "2F5496",
    "MaxDD": "C65911",
    "Turnover (ann.)": "7030A0",
    "TO (monthly %, GKX)": "7030A0",
    "D1": "BF8F00",
    "D10": "BF8F00",
}

DECILES_HEADERS = [
    "Market",
    "Weight",
    "Freq",
    "Config",
    "DH",
    *[f"D{i}" for i in range(1, 11)],
]
DECILES_FMT = {
    "Market": "text",
    "Weight": "text",
    "Freq": "text",
    "Config": "text",
    "DH": "pct",
    **{f"D{i}": "pct" for i in range(1, 11)},
}
DECILES_FILLS = {
    "Market": "4472C4",
    "Weight": "4472C4",
    "Freq": "4472C4",
    "Config": "4472C4",
    "DH": "548235",
    **{f"D{i}": "BF8F00" for i in range(1, 11)},
}


@dataclass(frozen=True)
class FreqSpec:
    name: str
    freq_dir: str
    horizon: int
    periods_per_year: int
    holding_months: float


FREQ_SPECS: dict[str, FreqSpec] = {
    "weekly": FreqSpec("weekly", "weekly", 5, 52, 0.25),
    "monthly": FreqSpec("monthly", "monthly", 20, 12, 1.0),
    "quarterly": FreqSpec("quarterly", "quarterly", 60, 4, 3.0),
}

WEIGHT_LABELS: dict[str, str] = {
    "equal": "EW",
    "float": "FloatCap",
    "total": "TotalCap",
}


def log(msg: str) -> None:
    print(msg, flush=True)


def read_h1_models(path: Path, sheet: str) -> dict[str, pd.Series]:
    raw = pd.read_excel(path, sheet_name=sheet, header=[0, 1])
    name_col = ("metric", "sig_rank")
    models: dict[str, pd.Series] = {}
    for _, row in raw.iterrows():
        model = row[name_col]
        if pd.isna(model) or str(model) in {"sig_rank", "nan", "t"}:
            continue
        models[str(model)] = row
    if not models:
        raise ValueError(f"no model rows in {path} sheet={sheet}")
    return models


def _num(row: pd.Series, metric: str, sig_rank: str) -> float:
    val = row[(metric, sig_rank)]
    if pd.isna(val):
        return float("nan")
    return float(val)


def engine_to_paper_monthly_pct(spec: FreqSpec, turnover_annualized: float) -> float:
    if not np.isfinite(turnover_annualized):
        return float("nan")
    per_rebalance = turnover_annualized / spec.periods_per_year
    return per_rebalance / spec.holding_months * 100.0


def discover_config_xlsx(
    backtest_root: Path,
    market: str,
    spec: FreqSpec,
    eval_horizon: int,
) -> list[tuple[str, Path]]:
    root = backtest_root / market / spec.freq_dir
    if not root.is_dir():
        return []
    out: list[tuple[str, Path]] = []
    for path in sorted(root.glob("I*_R*_h*.xlsx")):
        if path.name.startswith("direct_"):
            continue
        m = _CONFIG_RE.match(path.name)
        if m is None:
            continue
        if int(m.group("h")) != eval_horizon:
            continue
        out.append((m.group("tag"), path))
    return out


def _parse_config_sort(tag: str) -> tuple[int, int]:
    m = re.match(r"I(?P<i>\d+)_R(?P<r>\d+)$", tag, re.IGNORECASE)
    if m is None:
        return (999, 999)
    return int(m.group("i")), int(m.group("r"))


def build_row(
    *,
    market: str,
    weight: str,
    spec: FreqSpec,
    config: str,
    path: Path,
    eval_horizon: int,
) -> dict[str, object]:
    models = read_h1_models(path, weight)
    if config not in models:
        raise KeyError(f"{path} sheet={weight!r} missing config {config!r}")
    row = models[config]
    to_ann = _num(row, TO_METRIC, "DH")
    return {
        "Market": market,
        "Weight": WEIGHT_LABELS[weight],
        "Freq": spec.name,
        "Config": config,
        "Ret1": _num(row, ACT_RET_METRIC, "DH"),
        "IR": _num(row, IR_METRIC, "DH"),
        "MaxDD": _num(row, MDD_METRIC, "DH"),
        "Turnover (ann.)": to_ann,
        "TO (monthly %, GKX)": engine_to_paper_monthly_pct(spec, to_ann),
        "D1": _num(row, ACT_RET_METRIC, "D01"),
        "D10": _num(row, ACT_RET_METRIC, "D10"),
        "_sort_market": MARKETS.index(market) if market in MARKETS else 99,
        "_sort_freq": list(FREQ_SPECS).index(spec.name),
        "_sort_i": _parse_config_sort(config)[0],
        "_sort_r": _parse_config_sort(config)[1],
        "_eval_h": eval_horizon,
        "_source": str(path),
    }


def build_deciles_row(base: dict[str, object], path: Path, weight: str, config: str) -> dict[str, object]:
    models = read_h1_models(path, weight)
    row = models[config]
    out = {
        "Market": base["Market"],
        "Weight": base["Weight"],
        "Freq": base["Freq"],
        "Config": base["Config"],
        "DH": base["Ret1"],
    }
    for i in range(1, 11):
        out[f"D{i}"] = _num(row, ACT_RET_METRIC, f"D{i:02d}")
    return out


def sort_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    return sorted(
        rows,
        key=lambda r: (
            int(r["_sort_market"]),
            str(r["Weight"]),
            int(r["_sort_freq"]),
            int(r["_sort_i"]),
            int(r["_sort_r"]),
        ),
    )


def build_notes(weight: str, eval_horizon: int) -> pd.DataFrame:
    lines = [
        ("markets", ", ".join(MARKETS)),
        ("weight_sheet", weight),
        ("eval_horizon", f"H{eval_horizon}"),
        ("sheets", "baseline_presentation, baseline_deciles, top500_presentation, top500_deciles"),
        (
            "Ret1",
            f"DH {ACT_RET_METRIC} from step-04 cnn_baseline / cnn_top500 xlsx (H{eval_horizon} eval).",
        ),
        (
            "IR",
            f"DH {IR_METRIC} (= active Sharpe on DH spread). PSRET presentation uses IR, not raw Sharpe.",
        ),
        ("MaxDD", f"DH {MDD_METRIC}."),
        (
            "Turnover (ann.)",
            f"DH {TO_METRIC} from backtest engine (mean two-way turnover × periods/year).",
        ),
        (
            "TO (monthly %, GKX)",
            "Paper Table I/II monthly turnover % on H-L: per_rebalance / holding_months × 100, "
            "scaled from engine annualized turnover (GKX 2020 / Jiang–Kelly–Xiu).",
        ),
        ("D1 / D10", f"{ACT_RET_METRIC} for D01 / D10."),
        ("baseline_source", f"{BACKTEST_CNN_ROOT}/{{market}}/{{freq}}/I*_R*_h{eval_horizon}.xlsx"),
        ("top500_source", f"{BACKTEST_CNN_TOP500_ROOT}/{{market}}/{{freq}}/I*_R*_h{eval_horizon}.xlsx"),
    ]
    return pd.DataFrame([{"item": k, "value": v} for k, v in lines])


def _apply_sheet_style(
    ws,
    headers: list[str],
    rows: list[dict],
    fmt_map: dict[str, str],
    fill_map: dict[str, str],
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Border(*(Side(style="thin") for _ in range(4)))
    normal_font = Font(size=10)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = PatternFill("solid", fgColor=fill_map[h])
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = thin

    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(row=ri, column=ci)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            cell.font = normal_font
            fmt = fmt_map[h]
            if val is None or (isinstance(val, float) and not np.isfinite(val)):
                cell.value = ""
                continue
            if fmt == "text":
                cell.value = val
            elif fmt == "pct":
                cell.value = float(val)
                cell.number_format = "0.00%"
            elif fmt == "ir":
                cell.value = float(val)
                cell.number_format = "0.000"
            elif fmt == "turn":
                cell.value = float(val)
                cell.number_format = "0.00"
            else:
                cell.value = val

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.auto_filter.ref = ws.dimensions


def _write_raw_sheet(ws, raw: pd.DataFrame) -> None:
    for ci, col in enumerate(raw.columns, 1):
        ws.cell(row=1, column=ci, value=col)
    for ri, record in enumerate(raw.to_dict(orient="records"), 2):
        for ci, col in enumerate(raw.columns, 1):
            ws.cell(row=ri, column=ci, value=record[col])


def collect_portfolio_rows(
    kind: PortfolioKind,
    markets: tuple[str, ...],
    weight: str,
    freqs: tuple[str, ...],
    eval_horizon: int,
) -> tuple[list[dict], list[dict], pd.DataFrame]:
    backtest_root = PORTFOLIO_ROOTS[kind]
    pres_rows: list[dict[str, object]] = []
    for market in markets:
        for freq in freqs:
            spec = FREQ_SPECS[freq]
            for config, path in discover_config_xlsx(backtest_root, market, spec, eval_horizon):
                pres_rows.append(
                    build_row(
                        market=market,
                        weight=weight,
                        spec=spec,
                        config=config,
                        path=path,
                        eval_horizon=eval_horizon,
                    )
                )
    pres_rows = sort_rows(pres_rows)
    dec_rows = [
        build_deciles_row(r, Path(str(r["_source"])), weight, str(r["Config"])) for r in pres_rows
    ]
    raw = pd.DataFrame(pres_rows)
    public_cols = [c for c in PRESENTATION_HEADERS if c in raw.columns]
    raw = raw[public_cols + ["_source"]].copy()
    pres_public = [{k: r[k] for k in PRESENTATION_HEADERS} for r in pres_rows]
    return pres_public, dec_rows, raw


def write_workbook(
    path: Path,
    *,
    baseline_pres: list[dict],
    baseline_dec: list[dict],
    baseline_raw: pd.DataFrame,
    top500_pres: list[dict],
    top500_dec: list[dict],
    top500_raw: pd.DataFrame,
    notes: pd.DataFrame,
) -> Path:
    from openpyxl import Workbook

    path = Path(path)
    tmp = path.with_suffix(".tmp.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    ws_notes = wb.create_sheet("notes")
    for r_idx, row in enumerate(notes.itertuples(index=False), 1):
        ws_notes.cell(row=r_idx, column=1, value=row.item)
        ws_notes.cell(row=r_idx, column=2, value=row.value)

    for sheet_name, headers, rows, fmt_map, fill_map in (
        ("baseline_presentation", PRESENTATION_HEADERS, baseline_pres, PRESENTATION_FMT, PRESENTATION_FILLS),
        ("top500_presentation", PRESENTATION_HEADERS, top500_pres, PRESENTATION_FMT, PRESENTATION_FILLS),
        ("baseline_deciles", DECILES_HEADERS, baseline_dec, DECILES_FMT, DECILES_FILLS),
        ("top500_deciles", DECILES_HEADERS, top500_dec, DECILES_FMT, DECILES_FILLS),
    ):
        ws = wb.create_sheet(sheet_name)
        _apply_sheet_style(ws, headers, rows, fmt_map, fill_map)

    ws_raw_b = wb.create_sheet("raw_baseline")
    _write_raw_sheet(ws_raw_b, baseline_raw)
    ws_raw_t = wb.create_sheet("raw_top500")
    _write_raw_sheet(ws_raw_t, top500_raw)

    wb.save(tmp)
    tmp.replace(path)
    return path


def output_path(weight: str, eval_horizon: int) -> Path:
    return HERE / f"05_backtest_overview_{weight}_h{eval_horizon}.xlsx"


def run(weight: str, freqs: tuple[str, ...], eval_horizon: int) -> Path:
    baseline_pres, baseline_dec, baseline_raw = collect_portfolio_rows(
        "baseline", MARKETS, weight, freqs, eval_horizon
    )
    top500_pres, top500_dec, top500_raw = collect_portfolio_rows(
        "top500", MARKETS, weight, freqs, eval_horizon
    )
    if not baseline_pres:
        raise FileNotFoundError(f"no cnn_baseline xlsx for markets={MARKETS} freqs={freqs} h{eval_horizon}")
    if not top500_pres:
        raise FileNotFoundError(f"no cnn_top500 xlsx for markets={MARKETS} freqs={freqs} h{eval_horizon}")

    notes = build_notes(weight, eval_horizon)
    out = output_path(weight, eval_horizon)
    write_workbook(
        out,
        baseline_pres=baseline_pres,
        baseline_dec=baseline_dec,
        baseline_raw=baseline_raw,
        top500_pres=top500_pres,
        top500_dec=top500_dec,
        top500_raw=top500_raw,
        notes=notes,
    )
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="05 backtest overview (baseline + top500, US + CN)")
    parser.add_argument(
        "--weight",
        choices=tuple(WEIGHT_LABELS),
        default="equal",
        help="H1 xlsx sheet / weight scheme (default equal EW)",
    )
    parser.add_argument("--freq", choices=tuple(FREQ_SPECS), default=None)
    parser.add_argument("--eval-horizon", type=int, default=1, choices=(1, 2, 3))
    args = parser.parse_args()

    freqs: tuple[str, ...] = (args.freq,) if args.freq is not None else tuple(FREQ_SPECS)
    out = run(args.weight, freqs, args.eval_horizon)
    log(
        f"wrote {out} "
        f"(baseline + top500; US+CN; {len(freqs)} freqs × discovered configs per market)"
    )


if __name__ == "__main__":
    main()
